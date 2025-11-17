
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List, Dict, Any
from ..deps import get_db, get_current_user
from ..models import Vendor
from ..schemas import VendorIn, VendorOut
from openpyxl import load_workbook
import io

router = APIRouter(prefix="/vendors", tags=["vendors"])

# 台灣縣市 → 區域排序優先序（越小越前面）
NORTH = {"臺北市","台北市","新北市","基隆市","桃園市","新竹市","新竹縣","宜蘭縣"}
CENTRAL = {"苗栗縣","臺中市","台中市","彰化縣","彰化市","南投縣","雲林縣"}
SOUTH = {"嘉義市","嘉義縣","臺南市","台南市","高雄市","屏東縣"}
EAST = {"花蓮縣","臺東縣","台東縣"}
OUTER = {"澎湖縣","金門縣","連江縣"}  # 馬祖=連江縣

def region_rank(city: str) -> int:
    c = city or ""
    if c in NORTH: return 0
    if c in CENTRAL: return 1
    if c in SOUTH: return 2
    if c in EAST: return 3
    if c in OUTER: return 4
    return 5  # 未匹配的最後

def parse_city(addr: Optional[str]) -> str:
    """從地址抓出縣市關鍵詞（很寬鬆的比對）。"""
    if not addr: return ""
    s = addr.strip()
    # 常見縣市關鍵詞（可再擴充）
    keys = [
        "臺北市","台北市","新北市","基隆市","桃園市","新竹市","新竹縣","宜蘭縣",
        "苗栗縣","臺中市","台中市","彰化縣","南投縣","雲林縣",
        "嘉義市","嘉義縣","臺南市","台南市","高雄市","屏東縣",
        "花蓮縣","臺東縣","台東縣",
        "澎湖縣","金門縣","連江縣"
    ]
    for k in keys:
        if k in s:
            return k
    return ""


@router.get("", response_model=List[VendorOut])
def list_vendors(q: Optional[str] = None, db: Session = Depends(get_db), user=Depends(get_current_user)):
    query = db.query(Vendor)
    if q:
        like = f"%{q}%"
        query = query.filter(Vendor.name.like(like))
    items = query.all()

    # 依「區域 → 縣市 → 名稱」排序
    def sort_key(v: Vendor):
        city = parse_city(v.address)
        return (region_rank(city), city, v.name or "")
    items.sort(key=sort_key)

    # 轉為輸出
    out: List[VendorOut] = []
    for v in items:
        out.append(VendorOut(
            id=v.id, name=v.name, address=v.address, phone=v.phone,
            contact_person=v.contact_person, fax=v.fax, email=v.email, other=v.other
        ))
    return out

@router.post("", response_model=VendorOut)
def create_vendor(v: VendorIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if db.query(Vendor).filter(Vendor.name == v.name).first():
        raise HTTPException(400, detail="Vendor already exists")
    obj = Vendor(**v.model_dump())
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

@router.get("/{vid}", response_model=VendorOut)
def get_vendor(vid: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    obj = db.query(Vendor).filter(Vendor.id == vid).first()
    if not obj:
        raise HTTPException(404, detail="Vendor not found")
    return obj

@router.put("/{vid}", response_model=VendorOut)
def update_vendor(vid: int, payload: Dict[str, Any] = Body(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    obj = db.query(Vendor).filter(Vendor.id == vid).first()
    if not obj:
        raise HTTPException(404, detail="Vendor not found")
    for k in ["name","address","phone","contact_person","fax","email","other"]:
        if k in payload:
            setattr(obj, k, payload[k])
    db.commit(); db.refresh(obj)
    return obj

@router.delete("/{vid}")
def delete_vendor(vid: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    obj = db.query(Vendor).filter(Vendor.id == vid).first()
    if not obj:
        raise HTTPException(404, detail="Vendor not found")
    db.delete(obj); db.commit()
    return {"ok": True}

@router.post("/import_excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        def find_col(names):
            for i, h in enumerate(headers):
                if any(n in h for n in names):
                    return i
            return None
        name_idx = find_col(["廠商名稱","名稱","name","公司"])
        addr_idx = find_col(["地址","addr"])
        phone_idx = find_col(["電話","phone"])
        contact_idx = find_col(["聯絡人","contact"])
        fax_idx = find_col(["傳真","fax"])
        email_idx = find_col(["email","電子郵件","信箱"])
        other_idx = find_col(["其他","備註","remark"])
        if name_idx is None:
            raise HTTPException(400, detail="Excel 缺少「廠商名稱」欄")
        cnt = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            name = str(r[name_idx]).strip() if r[name_idx] is not None else ""
            if not name:
                continue
            if db.query(Vendor).filter(Vendor.name==name).first():
                continue
            obj = Vendor(
                name=name,
                address=(str(r[addr_idx]).strip() if addr_idx is not None and r[addr_idx] is not None else None),
                phone=(str(r[phone_idx]).strip() if phone_idx is not None and r[phone_idx] is not None else None),
                contact_person=(str(r[contact_idx]).strip() if contact_idx is not None and r[contact_idx] is not None else None),
                fax=(str(r[fax_idx]).strip() if fax_idx is not None and r[fax_idx] is not None else None),
                email=(str(r[email_idx]).strip() if email_idx is not None and r[email_idx] is not None else None),
                other=(str(r[other_idx]).strip() if other_idx is not None and r[other_idx] is not None else None),
            )
            db.add(obj); cnt += 1
        db.commit()
        return {"detail": f"匯入完成，共新增 {cnt} 筆"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, detail=f"匯入失敗：{e}")
